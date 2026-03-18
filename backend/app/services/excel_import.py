from typing import List, Tuple
from io import BytesIO, StringIO
import csv
import openpyxl
from pydantic import BaseModel, ValidationError

from app.schemas.resident import ResidentRow

# Expected column names (Hebrew or English)
COLUMN_ALIASES = {
    "identity_number": ["מספר זהות", "תעודת זהות", "id", "identity_number"],
    "gender": ["מין", "gender"],
    "city": ["ישוב", "יישוב", "city"],
    "street": ["רחוב", "street"],
    "house_number": ["מס' בית", "מס בית", "house_number", "house number"],
    "apartment": ["דירה", "apartment"],
    "age": ["גיל", "age"],
    "first_name": ["שם פרטי", "first_name", "first name", "שם"],
    "last_name": ["שם משפחה", "last_name", "last name", "משפחה"],
    "phone": ["טלפון נייד", "טלפון", "phone", "phone number", "mobile phone"],
    "home_phone": ["טלפון בבית", "טלפון בית", "home_phone", "home phone"],
    "notes": ["הערות", "notes", "הערות מקדימות"],
}


def _normalize_header(h: str) -> str:
    return (h or "").strip().lower()


def _find_column_index(headers: List[str]) -> dict:
    """Map field name to 0-based column index from first row."""
    result = {}
    for idx, cell in enumerate(headers):
        val = _normalize_header(str(cell) if cell is not None else "")
        for field, aliases in COLUMN_ALIASES.items():
            if val in [a.strip().lower() for a in aliases]:
                result[field] = idx
                break
    return result


def _cell_to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _compose_address(city: str, street: str, house_number: str, apartment: str) -> str:
    parts = [part for part in [city, street] if part]
    if house_number:
        street_part = f"{street} {house_number}".strip() if street else house_number
        parts = [part for part in [city, street_part] if part]
    address = ", ".join(parts)
    if apartment:
        address = f"{address} דירה {apartment}".strip()
    return address.strip(", ")


def _parse_age(raw_value) -> int | None:
    text = _cell_to_text(raw_value)
    if not text:
        return None
    try:
        return int(text)
    except ValueError as exc:
        raise ValueError("גיל חייב להיות מספר שלם") from exc


def parse_residents_excel(content: bytes) -> Tuple[List[ResidentRow], List[str]]:
    """Parse Excel file, return list of valid ResidentRow and list of error messages."""
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if not ws:
        return [], ["גיליון ריק"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], ["אין שורות"]
    headers = list(rows[0])
    col = _find_column_index(headers)
    if (
        "identity_number" not in col
        or "first_name" not in col
        or "last_name" not in col
    ):
        return [], ["חסרות עמודות נדרשות: מספר זהות, שם משפחה, שם פרטי"]
    out: List[ResidentRow] = []
    errors: List[str] = []
    for i, row in enumerate(rows[1:], start=2):
        if not any(cell is not None for cell in row):
            continue
        identity_number = _cell_to_text(
            row[col["identity_number"]] if col["identity_number"] < len(row) else None
        )
        first_name = _cell_to_text(
            row[col["first_name"]] if col["first_name"] < len(row) else None
        )
        last_name = _cell_to_text(
            row[col["last_name"]] if col["last_name"] < len(row) else None
        )
        gender = _cell_to_text(row[col["gender"]] if "gender" in col and col["gender"] < len(row) else None) or None
        city = _cell_to_text(row[col["city"]] if "city" in col and col["city"] < len(row) else None)
        street = _cell_to_text(row[col["street"]] if "street" in col and col["street"] < len(row) else None)
        house_number = _cell_to_text(
            row[col["house_number"]] if "house_number" in col and col["house_number"] < len(row) else None
        )
        apartment = _cell_to_text(
            row[col["apartment"]] if "apartment" in col and col["apartment"] < len(row) else None
        )
        address = _compose_address(city, street, house_number, apartment)
        phone = _cell_to_text(row[col["phone"]] if "phone" in col and col["phone"] < len(row) else None) or None
        home_phone = _cell_to_text(
            row[col["home_phone"]] if "home_phone" in col and col["home_phone"] < len(row) else None
        ) or None
        notes = _cell_to_text(row[col["notes"]] if "notes" in col and col["notes"] < len(row) else None) or None
        try:
            out.append(
                ResidentRow(
                    identity_number=identity_number,
                    gender=gender,
                    city=city or None,
                    street=street or None,
                    house_number=house_number or None,
                    apartment=apartment or None,
                    age=_parse_age(
                        row[col["age"]] if "age" in col and col["age"] < len(row) else None
                    ),
                    first_name=first_name,
                    last_name=last_name,
                    address=address,
                    phone=phone or None,
                    home_phone=home_phone or None,
                    notes=notes or None,
                )
            )
        except (ValidationError, ValueError) as e:
            errors.append(f"שורה {i}: {e}")
    wb.close()
    return out, errors


def parse_residents_csv(content: bytes) -> Tuple[List[ResidentRow], List[str]]:
    """Parse CSV file (UTF-8), return list of valid ResidentRow and list of error messages."""
    try:
        text = content.decode("utf-8-sig").strip()
    except UnicodeDecodeError:
        return [], ["קובץ לא ב-UTF-8"]
    if not text:
        return [], ["קובץ ריק"]
    reader = csv.reader(StringIO(text))
    rows = list(reader)
    if not rows:
        return [], ["אין שורות"]
    headers = rows[0]
    col = _find_column_index(headers)
    if (
        "identity_number" not in col
        or "first_name" not in col
        or "last_name" not in col
    ):
        return [], ["חסרות עמודות נדרשות: מספר זהות, שם משפחה, שם פרטי"]
    out: List[ResidentRow] = []
    errors: List[str] = []
    for i, row in enumerate(rows[1:], start=2):
        if not row or not any(cell and str(cell).strip() for cell in row):
            continue
        identity_number = _cell_to_text(
            row[col["identity_number"]] if col["identity_number"] < len(row) else ""
        )
        first_name = _cell_to_text(
            row[col["first_name"]] if col["first_name"] < len(row) else ""
        )
        last_name = _cell_to_text(
            row[col["last_name"]] if col["last_name"] < len(row) else ""
        )
        gender = _cell_to_text(row[col["gender"]] if "gender" in col and col["gender"] < len(row) else "") or None
        city = _cell_to_text(row[col["city"]] if "city" in col and col["city"] < len(row) else "")
        street = _cell_to_text(row[col["street"]] if "street" in col and col["street"] < len(row) else "")
        house_number = _cell_to_text(
            row[col["house_number"]] if "house_number" in col and col["house_number"] < len(row) else ""
        )
        apartment = _cell_to_text(
            row[col["apartment"]] if "apartment" in col and col["apartment"] < len(row) else ""
        )
        address = _compose_address(city, street, house_number, apartment)
        phone = _cell_to_text(row[col["phone"]] if "phone" in col and col["phone"] < len(row) else "") or None
        home_phone = _cell_to_text(
            row[col["home_phone"]] if "home_phone" in col and col["home_phone"] < len(row) else ""
        ) or None
        notes = _cell_to_text(row[col["notes"]] if "notes" in col and col["notes"] < len(row) else "") or None
        try:
            out.append(
                ResidentRow(
                    identity_number=identity_number,
                    gender=gender,
                    city=city or None,
                    street=street or None,
                    house_number=house_number or None,
                    apartment=apartment or None,
                    age=_parse_age(
                        row[col["age"]] if "age" in col and col["age"] < len(row) else None
                    ),
                    first_name=first_name,
                    last_name=last_name,
                    address=address,
                    phone=phone or None,
                    home_phone=home_phone or None,
                    notes=notes or None,
                )
            )
        except (ValidationError, ValueError) as e:
            errors.append(f"שורה {i}: {e}")
    return out, errors


def parse_residents_file(content: bytes, filename: str) -> Tuple[List[ResidentRow], List[str]]:
    """Dispatch to Excel or CSV parser based on file extension."""
    fn = (filename or "").lower()
    if fn.endswith(".csv"):
        return parse_residents_csv(content)
    if fn.endswith(".xlsx") or fn.endswith(".xls"):
        return parse_residents_excel(content)
    return [], ["סוג קובץ לא נתמך. השתמש ב־CSV או Excel (xlsx)."]


VOLUNTEER_COLUMN_ALIASES = {
    "first_name": ["first_name", "first name", "שם פרטי", "שם"],
    "last_name": ["last_name", "last name", "שם משפחה", "משפחה"],
    "phone": ["phone", "mobile phone", "טלפון", "טלפון נייד"],
    "group_tag": ["group_tag", "group", "קבוצה", "התמחות"],
    "living_area": ["living_area", "area", "אזור", "אזור מגורים"],
}


class VolunteerImportRow(BaseModel):
    first_name: str
    last_name: str
    phone: str
    group_tag: str | None = None
    living_area: str | None = None


def _find_volunteer_column_index(headers: List[str]) -> dict:
    result = {}
    for idx, cell in enumerate(headers):
        val = _normalize_header(str(cell) if cell is not None else "")
        for field, aliases in VOLUNTEER_COLUMN_ALIASES.items():
            if val in [a.strip().lower() for a in aliases]:
                result[field] = idx
                break
    return result


def parse_volunteers_excel(content: bytes) -> Tuple[List[VolunteerImportRow], List[str]]:
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if not ws:
        return [], ["גיליון ריק"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], ["אין שורות"]
    headers = list(rows[0])
    col = _find_volunteer_column_index(headers)
    if "first_name" not in col or "last_name" not in col or "phone" not in col:
        return [], ["חסרות עמודות נדרשות: first_name, last_name, phone"]

    out: List[VolunteerImportRow] = []
    errors: List[str] = []
    for i, row in enumerate(rows[1:], start=2):
        if not any(cell is not None for cell in row):
            continue
        try:
            out.append(
                VolunteerImportRow(
                    first_name=_cell_to_text(row[col["first_name"]] if col["first_name"] < len(row) else ""),
                    last_name=_cell_to_text(row[col["last_name"]] if col["last_name"] < len(row) else ""),
                    phone=_cell_to_text(row[col["phone"]] if col["phone"] < len(row) else ""),
                    group_tag=_cell_to_text(row[col["group_tag"]] if "group_tag" in col and col["group_tag"] < len(row) else "") or None,
                    living_area=_cell_to_text(row[col["living_area"]] if "living_area" in col and col["living_area"] < len(row) else "") or None,
                )
            )
        except ValidationError as e:
            errors.append(f"שורה {i}: {e}")
    wb.close()
    return out, errors


def parse_volunteers_csv(content: bytes) -> Tuple[List[VolunteerImportRow], List[str]]:
    try:
        text = content.decode("utf-8-sig").strip()
    except UnicodeDecodeError:
        return [], ["קובץ לא ב-UTF-8"]
    if not text:
        return [], ["קובץ ריק"]
    rows = list(csv.reader(StringIO(text)))
    if not rows:
        return [], ["אין שורות"]
    col = _find_volunteer_column_index(rows[0])
    if "first_name" not in col or "last_name" not in col or "phone" not in col:
        return [], ["חסרות עמודות נדרשות: first_name, last_name, phone"]

    out: List[VolunteerImportRow] = []
    errors: List[str] = []
    for i, row in enumerate(rows[1:], start=2):
        if not row or not any(cell and str(cell).strip() for cell in row):
            continue
        try:
            out.append(
                VolunteerImportRow(
                    first_name=_cell_to_text(row[col["first_name"]] if col["first_name"] < len(row) else ""),
                    last_name=_cell_to_text(row[col["last_name"]] if col["last_name"] < len(row) else ""),
                    phone=_cell_to_text(row[col["phone"]] if col["phone"] < len(row) else ""),
                    group_tag=_cell_to_text(row[col["group_tag"]] if "group_tag" in col and col["group_tag"] < len(row) else "") or None,
                    living_area=_cell_to_text(row[col["living_area"]] if "living_area" in col and col["living_area"] < len(row) else "") or None,
                )
            )
        except ValidationError as e:
            errors.append(f"שורה {i}: {e}")
    return out, errors


def parse_volunteers_file(content: bytes, filename: str) -> Tuple[List[VolunteerImportRow], List[str]]:
    fn = (filename or "").lower()
    if fn.endswith(".csv"):
        return parse_volunteers_csv(content)
    if fn.endswith(".xlsx") or fn.endswith(".xls"):
        return parse_volunteers_excel(content)
    return [], ["סוג קובץ לא נתמך. השתמש ב־CSV או Excel (xlsx)."]
