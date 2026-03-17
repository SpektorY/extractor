from datetime import datetime, timezone
from io import BytesIO

import openpyxl

from app.models import Event, Resident, ResidentStatus, ResidentSource
from app.services.excel_export import export_event_to_excel
from app.services.excel_import import parse_residents_excel


def build_template_workbook(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "גיליון1"
    ws.append(
        [
            "מספר זהות",
            "שם משפחה",
            "שם פרטי",
            "מין",
            "ישוב",
            "רחוב",
            "מס' בית",
            "דירה",
            "גיל",
            "טלפון נייד",
            "טלפון בבית",
        ]
    )
    for row in rows:
        ws.append(row)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_parse_residents_excel_matches_exact_template():
    content = build_template_workbook(
        [
            ["17491952", "אברמוב", "לזר", "זכר", "אלעד", "שמעון הצדיק", "60", "15", 50, "0501111111", "039999999"],
        ]
    )

    rows, errors = parse_residents_excel(content)

    assert errors == []
    assert len(rows) == 1
    row = rows[0]
    assert row.identity_number == "17491952"
    assert row.last_name == "אברמוב"
    assert row.first_name == "לזר"
    assert row.gender == "זכר"
    assert row.city == "אלעד"
    assert row.street == "שמעון הצדיק"
    assert row.house_number == "60"
    assert row.apartment == "15"
    assert row.age == 50
    assert row.phone == "0501111111"
    assert row.home_phone == "039999999"
    assert row.address == "אלעד, שמעון הצדיק 60 דירה 15"


def test_export_event_to_excel_uses_exact_template_columns_and_event_log():
    resident = Resident(
        identity_number="17491952",
        last_name="אברמוב",
        first_name="לזר",
        gender="זכר",
        city="אלעד",
        street="שמעון הצדיק",
        house_number="60",
        apartment="15",
        age=50,
        address="אלעד, שמעון הצדיק 60 דירה 15",
        phone="0501111111",
        home_phone="039999999",
        notes="קשיש",
        status=ResidentStatus.UNCHECKED,
        volunteer_notes="לא נוצר קשר",
        source=ResidentSource.UPLOADED.value,
        updated_at=datetime(2026, 3, 17, 9, 30, tzinfo=timezone.utc),
    )

    buffer = export_event_to_excel(
        Event(name="בדיקה", address="אלעד", description=None),
        [resident],
        [
            {
                "created_at": datetime(2026, 3, 17, 9, 45, tzinfo=timezone.utc),
                "author_type": "admin",
                "author_name": "מנהל",
                "message": "מתחילים לעבור דירה דירה",
            }
        ],
    )

    wb = openpyxl.load_workbook(buffer)
    ws = wb["גיליון1"]
    log_ws = wb["יומן אירוע"]

    header = [cell.value for cell in ws[1]]
    first_row = [cell.value for cell in ws[2]]
    log_header = [cell.value for cell in log_ws[1]]
    log_row = [cell.value for cell in log_ws[2]]

    assert wb.sheetnames == ["גיליון1", "יומן אירוע"]
    assert header == [
        "מספר זהות",
        "שם משפחה",
        "שם פרטי",
        "מין",
        "ישוב",
        "רחוב",
        "מס' בית",
        "דירה",
        "גיל",
        "טלפון נייד",
        "טלפון בבית",
        "הערות מקדימות",
        "סטטוס נוכחי",
        "הערות מתנדב",
        "מקור",
        "עודכן לאחרונה",
    ]
    assert first_row == [
        "17491952",
        "אברמוב",
        "לזר",
        "זכר",
        "אלעד",
        "שמעון הצדיק",
        "60",
        "15",
        50,
        "0501111111",
        "039999999",
        "קשיש",
        "unchecked",
        "לא נוצר קשר",
        "uploaded",
        "2026-03-17T09:30:00+00:00",
    ]
    assert log_header == ["תאריך", "סוג כותב", "שם כותב", "הודעה"]
    assert log_row == [
        "2026-03-17T09:45:00+00:00",
        "admin",
        "מנהל",
        "מתחילים לעבור דירה דירה",
    ]
